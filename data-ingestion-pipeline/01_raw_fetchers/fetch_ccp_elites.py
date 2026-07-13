"""
Raw CCP elite leadership fetcher — Sine CPC Elite Leadership Database.

Downloads the source xlsx and stores **every sheet raw** (CC / PB / PSC members,
full 1945–2022 coverage) to ``02_inputs/ccp_elites/<sheet>``, append-only. Column
names are taken from each sheet's header row — no reshaping.

Counterpart to ``backend/runners/import_ccp_elites.py`` (parses into 3 cmm.db
tables). Reuses the same source URL and the cached xlsx under ``data/raw/``.
"""

from __future__ import annotations

import argparse
import logging
import sys
import urllib.request
from pathlib import Path

import openpyxl
import pandas as pd

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE))
sys.path.insert(0, str(_HERE.parents[1]))

import _raw_store as store  # noqa: E402
from backend.storage import RAW_DIR  # noqa: E402

log = logging.getLogger("raw_ccp")
SOURCE = "ccp_elites"
SOURCE_URL = "https://cpcleadershipdata.pages.dev/CPC_Elite_Leadership_Database.xlsx"
XLSX = RAW_DIR / "cpc_elite_leadership.xlsx"


def _sheet_to_df(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    header = [str(c) if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
    return pd.DataFrame(rows[1:], columns=header)


def run(run_id: str | None = None, *, force_download: bool = False) -> None:
    run_id = run_id or store.new_run_id()
    datasets: dict[str, int] = {}
    try:
        if force_download or not XLSX.exists():
            XLSX.parent.mkdir(parents=True, exist_ok=True)
            log.info("Downloading %s", SOURCE_URL)
            urllib.request.urlretrieve(SOURCE_URL, XLSX)
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            df = _sheet_to_df(wb[sheet])
            ds = sheet.lower().replace(" ", "_")
            store.append(SOURCE, ds, df, run_id=run_id, endpoint=SOURCE_URL)
            datasets[ds] = len(df)
            log.info("  %-14s %5d rows", sheet, len(df))
        store.write_manifest(SOURCE, status="ok", datasets=datasets, run_id=run_id)
    except Exception as e:
        store.write_manifest(SOURCE, status="error", datasets=datasets,
                             run_id=run_id, error=str(e))
        raise


def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s",
                        datefmt="%H:%M:%S")
    ap = argparse.ArgumentParser(description="Raw CCP elites fetcher")
    ap.add_argument("--force-download", action="store_true")
    args = ap.parse_args()
    run(force_download=args.force_download)


if __name__ == "__main__":
    main()
