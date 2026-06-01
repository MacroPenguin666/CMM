"""
Fetch China economic data from the Bruegel China Economic Database.

Downloads Excel files from https://china-dashboard.herokuapp.com/data/,
parses macro, financial, and structural indicators, and stores them in
data/feeds.db in the `bruegel_series` and `bruegel_meta` tables.

Uses If-Modified-Since headers to avoid re-downloading unchanged files
(data updates roughly monthly).

Usage:
    python bruegel.py                 # fetch all available data
    python bruegel.py --show          # show latest stored snapshots
    python bruegel.py --json          # export as JSON
"""

import argparse
import json
import logging
import re
import sqlite3
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import requests
from openpyxl import load_workbook

from policy_monitor.storage import DB_DIR, DB_PATH

log = logging.getLogger("bruegel")

BASE_URL = "https://china-dashboard.herokuapp.com/data"

# Files to download, in priority order
SOURCE_FILES = [
    "China High Frequency v2.xlsx",
    "Financial_Indicators.xlsx",
    "China Low Frequency v2.xlsx",
    "Provincial Profiles.xlsx",
]

# Files with pivot layout (regions as rows, years as columns)
PIVOT_FILES = {"Provincial Profiles.xlsx"}

# Sheets with non-standard layouts to skip
SKIP_SHEETS = {
    "Real Estate - Home Sales",  # year-column pivot layout
}

# Sheet name prefix → category mapping
# Uses prefix matching so "Consumption - Box Office" matches "Consumption"
CATEGORY_PREFIXES = [
    # Order matters: longer/more specific prefixes first
    ("Real Estate", "real_estate"),
    ("Consumption", "consumption"),
    ("Retail Sales", "consumption"),
    ("Production", "production"),
    ("Industrial", "production"),
    ("Steel", "production"),
    ("Investment", "investment"),
    ("Export", "trade"),
    ("FDI", "external"),
    ("PMI", "macro"),
    ("Caixin PMI", "macro"),
    ("Inflation", "macro"),
    ("M2", "macro"),
    ("SHIBOR", "financial"),
    ("Govt Bonds Yield", "financial"),
    ("LPR", "financial"),
    ("Market lending rate", "financial"),
    ("Exchange Rates", "financial"),
    ("CFETS RMB Index", "financial"),
    ("RMB Index", "financial"),
    ("Bonds", "financial"),
    ("Stock and Bonds", "financial"),
    ("Monetary Policy", "monetary"),
    ("Stacked PBOC Loan", "monetary"),
    ("PBOC Loan GDP", "monetary"),
    ("Shadow Banking", "monetary"),
    ("TSF", "finance"),
    ("RRR", "finance"),
    ("Bonds GDP", "finance"),
    ("Stock GDP", "finance"),
    ("Total Loan", "banking"),
    ("Property Loan", "banking"),
    ("Corporate Loan", "banking"),
    ("Mortgage Loan", "banking"),
    ("Household Loans", "banking"),
    ("Commercial Loans", "banking"),
    ("Assets", "banking"),
    ("NPL Ratio", "banking"),
    ("US Securities", "external"),
    ("Gold", "external"),
    ("Current Account", "external"),
    ("Balance of Payment", "external"),
    ("External", "external"),
    ("LGFV", "fiscal"),
    ("Fiscal Deficit", "fiscal"),
    ("Debt by Sector", "fiscal"),
    ("Public Debt", "fiscal"),
    ("Household Debt", "fiscal"),
    ("Energy", "structural"),
    ("Labour", "structural"),
    ("EU-China", "eu_china"),
    ("China Corporate Monitor", "eu_china"),
    ("Green Sector", "green"),
    ("Financial", "financial"),
    ("Provincial", "structural"),
    ("Urbanization", "structural"),
    ("Innovation", "structural"),
    ("Industrial Policy", "structural"),
]


def _get_category(sheet_name: str) -> str:
    """Map a sheet name to a category using prefix matching."""
    for prefix, category in CATEGORY_PREFIXES:
        if sheet_name.startswith(prefix) or sheet_name == prefix:
            return category
    return "other"

# ---------------------------------------------------------------------------
# DB schema
# ---------------------------------------------------------------------------
BRUEGEL_SCHEMA = """
CREATE TABLE IF NOT EXISTS bruegel_series (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    indicator   TEXT NOT NULL,
    category    TEXT NOT NULL,
    source_file TEXT NOT NULL,
    date        TEXT,
    value       REAL,
    unit        TEXT,
    fetched_at  TEXT NOT NULL,
    UNIQUE(indicator, date)
);

CREATE TABLE IF NOT EXISTS bruegel_provincial (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    province    TEXT NOT NULL,
    indicator   TEXT NOT NULL,
    year        INTEGER NOT NULL,
    value       REAL,
    unit        TEXT,
    fetched_at  TEXT NOT NULL,
    UNIQUE(province, indicator, year)
);

CREATE TABLE IF NOT EXISTS bruegel_meta (
    source_file     TEXT PRIMARY KEY,
    last_modified   TEXT,
    etag            TEXT,
    last_fetched    TEXT NOT NULL,
    row_count       INTEGER DEFAULT 0
);
"""


def get_bruegel_db() -> sqlite3.Connection:
    DB_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))
    conn.executescript(BRUEGEL_SCHEMA)
    return conn


# ---------------------------------------------------------------------------
# Download with caching
# ---------------------------------------------------------------------------

def _download_if_modified(filename: str, conn: sqlite3.Connection) -> bytes | None:
    """Download file only if modified since last fetch. Returns bytes or None."""
    url = f"{BASE_URL}/{requests.utils.quote(filename)}"

    # Check stored Last-Modified
    row = conn.execute(
        "SELECT last_modified, etag FROM bruegel_meta WHERE source_file = ?",
        (filename,),
    ).fetchone()

    headers = {"User-Agent": "ChinaPolicyMonitor/1.0"}
    if row:
        if row[0]:
            headers["If-Modified-Since"] = row[0]
        if row[1]:
            headers["If-None-Match"] = row[1]

    try:
        resp = requests.get(url, headers=headers, timeout=60)
    except requests.RequestException as e:
        log.warning(f"  FAIL download {filename}: {e}")
        return None

    if resp.status_code == 304:
        log.info(f"  SKIP {filename} (not modified)")
        return None

    if resp.status_code != 200:
        log.warning(f"  FAIL {filename}: HTTP {resp.status_code}")
        return None

    # Update meta
    now = datetime.now(timezone.utc).isoformat()
    conn.execute(
        "INSERT OR REPLACE INTO bruegel_meta (source_file, last_modified, etag, last_fetched) "
        "VALUES (?, ?, ?, ?)",
        (filename, resp.headers.get("Last-Modified"), resp.headers.get("ETag"), now),
    )
    conn.commit()

    log.info(f"  OK   {filename} ({len(resp.content)} bytes)")
    return resp.content


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _normalize_indicator(sheet_name: str, col_header: str) -> str:
    """Create a clean indicator name: BRU_{sheet}_{column}."""
    # Clean sheet name
    sheet = re.sub(r'[^\w\s-]', '', sheet_name).strip()
    sheet = re.sub(r'[\s-]+', '_', sheet)

    # Clean column header
    col = str(col_header).strip()
    col = re.sub(r'[^\w\s%/-]', '', col)
    col = re.sub(r'[\s/]+', '_', col)
    col = col.strip('_')

    return f"BRU_{sheet}_{col}"


def _parse_date(val) -> str | None:
    """Parse various date formats from Bruegel Excel sheets."""
    if val is None:
        return None

    # datetime objects (most common)
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    val_str = str(val).strip()
    if not val_str:
        return None

    # "YYYY-MM" format
    if re.match(r'^\d{4}-\d{2}$', val_str):
        return val_str + "-01"

    # "DD/MM/YYYY" format
    if re.match(r'^\d{2}/\d{2}/\d{4}$', val_str):
        try:
            return datetime.strptime(val_str, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            return None

    # Integer year (2000-2099)
    if isinstance(val, (int, float)) and 2000 <= val <= 2099:
        return f"{int(val)}-01-01"

    # String year
    if re.match(r'^\d{4}$', val_str):
        return f"{val_str}-01-01"

    return None


def _parse_excel(data: bytes, source_file: str) -> tuple[list, list]:
    """Parse a Bruegel Excel file into rows and snapshots."""
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    all_rows = []
    all_snapshots = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        try:
            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            if len(rows_data) < 2:
                continue

            # Handle two-row headers (e.g., EU-China Trade Deficit)
            header_row = rows_data[0]
            data_start = 1

            # Detect if row 1 is a title (very few non-None cells, row 2 has more)
            non_none_r1 = sum(1 for v in header_row if v is not None)
            if len(rows_data) > 2:
                non_none_r2 = sum(1 for v in rows_data[1] if v is not None)
                if non_none_r1 <= 2 and non_none_r2 > non_none_r1:
                    header_row = rows_data[1]
                    data_start = 2

            # Build column map: index → (indicator_name, header)
            columns = {}
            for i, hdr in enumerate(header_row):
                if i == 0:
                    continue  # date column
                if hdr is None:
                    continue
                hdr_str = str(hdr).strip()
                if not hdr_str:
                    continue
                indicator = _normalize_indicator(sheet_name, hdr_str)
                columns[i] = (indicator, hdr_str)

            if not columns:
                continue

            category = _get_category(sheet_name)

            # Track last valid value per indicator for snapshots
            last_valid = {}  # indicator → (date, value)

            for row in rows_data[data_start:]:
                date_str = _parse_date(row[0] if row else None)
                if date_str is None:
                    continue

                # Check if entire row (besides date) is None
                has_data = False
                for col_idx, (indicator, _hdr) in columns.items():
                    if col_idx >= len(row):
                        continue
                    val = row[col_idx]

                    # Handle #N/A strings and empty values
                    if val is None or str(val).strip() in ('', '#N/A', '#N/A!', '#REF!', '#VALUE!'):
                        continue

                    try:
                        fval = float(val)
                    except (ValueError, TypeError):
                        continue

                    has_data = True
                    all_rows.append((indicator, category, source_file, date_str, fval, ""))
                    last_valid[indicator] = (date_str, fval)

            # Generate snapshots from last valid values
            for col_idx, (indicator, hdr_str) in columns.items():
                if indicator in last_valid:
                    date_str, value = last_valid[indicator]
                    all_snapshots.append({
                        "indicator": indicator,
                        "category": category,
                        "latest_value": value,
                        "change": None,
                        "unit": "",
                        "data_date": date_str,
                    })

        except Exception as e:
            log.warning(f"  WARN sheet '{sheet_name}' in {source_file}: {type(e).__name__}: {str(e)[:80]}")
            continue

    wb.close()
    return all_rows, all_snapshots


# ---------------------------------------------------------------------------
# Provincial pivot parser
# ---------------------------------------------------------------------------

# Sheet title → (indicator_name, unit)
PROVINCIAL_INDICATORS = {
    "GDP": ("GDP", "100M_yuan"),
    "Population": ("Population", "10K_people"),
    "Export Value": ("Exports", "1000_USD"),
}


def _parse_provincial(data: bytes, source_file: str) -> list:
    """Parse a pivot-layout file (regions as rows, years as columns).
    Returns list of (province, indicator, year, value, unit).
    """
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            if len(rows_data) < 3:
                continue

            # Row 0 is the title/description, row 1 is the header with years
            # Find the header row (the one with year integers)
            header_row_idx = None
            for idx in range(min(3, len(rows_data))):
                row = rows_data[idx]
                year_count = sum(1 for v in row[1:] if isinstance(v, (int, float)) and 2000 <= v <= 2099)
                if year_count >= 3:
                    header_row_idx = idx
                    break
            if header_row_idx is None:
                continue

            header = rows_data[header_row_idx]
            years = []
            for i, v in enumerate(header):
                if i == 0:
                    continue
                if isinstance(v, (int, float)) and 2000 <= v <= 2099:
                    years.append((i, int(v)))

            if not years:
                continue

            # Determine indicator and unit from sheet name or PROVINCIAL_INDICATORS
            ind_info = PROVINCIAL_INDICATORS.get(sheet_name)
            if ind_info:
                indicator, unit = ind_info
            else:
                indicator = re.sub(r'[^\w\s-]', '', sheet_name).strip().replace(' ', '_')
                unit = ""

            # Parse data rows (regions)
            for row in rows_data[header_row_idx + 1:]:
                province = row[0]
                if not province or not str(province).strip():
                    continue
                province = str(province).strip()

                for col_idx, year in years:
                    if col_idx >= len(row):
                        continue
                    val = row[col_idx]
                    if val is None or str(val).strip() in ('', '#N/A', '#N/A!'):
                        continue
                    try:
                        fval = float(val)
                    except (ValueError, TypeError):
                        continue
                    all_rows.append((province, indicator, year, fval, unit))

        except Exception as e:
            log.warning(f"  WARN provincial sheet '{sheet_name}': {type(e).__name__}: {str(e)[:80]}")
            continue

    wb.close()
    return all_rows


def store_provincial_data(conn: sqlite3.Connection, rows: list) -> int:
    """Store provincial pivot data. Returns count of new rows inserted."""
    now = datetime.now(timezone.utc).isoformat()
    inserted = 0
    for province, indicator, year, value, unit in rows:
        try:
            conn.execute(
                "INSERT OR IGNORE INTO bruegel_provincial "
                "(province, indicator, year, value, unit, fetched_at) "
                "VALUES (?,?,?,?,?,?)",
                (province, indicator, year, value, unit, now),
            )
            inserted += 1
        except sqlite3.IntegrityError:
            pass
    conn.commit()
    return inserted


def get_provincial_data(conn: sqlite3.Connection, indicator: str = "", year: int = 0) -> list[dict]:
    """Return provincial data, optionally filtered by indicator and/or year."""
    query = "SELECT province, indicator, year, value, unit FROM bruegel_provincial WHERE 1=1"
    params = []
    if indicator:
        query += " AND indicator = ?"
        params.append(indicator)
    if year:
        query += " AND year = ?"
        params.append(year)
    query += " ORDER BY province, year"
    cur = conn.execute(query, params)
    return [{"province": r[0], "indicator": r[1], "year": r[2], "value": r[3], "unit": r[4]}
            for r in cur.fetchall()]


def get_provincial_indicators(conn: sqlite3.Connection) -> list[dict]:
    """Return available provincial indicators and their year ranges."""
    cur = conn.execute(
        "SELECT indicator, unit, MIN(year), MAX(year), COUNT(DISTINCT province) "
        "FROM bruegel_provincial GROUP BY indicator ORDER BY indicator"
    )
    return [{"indicator": r[0], "unit": r[1], "min_year": r[2], "max_year": r[3], "provinces": r[4]}
            for r in cur.fetchall()]


# ---------------------------------------------------------------------------
# Store results
# ---------------------------------------------------------------------------

def store_bruegel_data(conn: sqlite3.Connection, rows: list, snapshots: list) -> int:
    """Store parsed Bruegel data. Returns count of new rows inserted."""
    now = datetime.now(timezone.utc).isoformat()
    inserted = 0
    for indicator, category, source_file, date, value, unit in rows:
        try:
            conn.execute(
                "INSERT OR IGNORE INTO bruegel_series "
                "(indicator, category, source_file, date, value, unit, fetched_at) "
                "VALUES (?,?,?,?,?,?,?)",
                (indicator, category, source_file, date, value, unit, now),
            )
            inserted += 1
        except sqlite3.IntegrityError:
            pass

    conn.commit()

    # Update row counts in meta
    for sf in set(r[2] for r in rows) if rows else []:
        count = conn.execute(
            "SELECT COUNT(*) FROM bruegel_series WHERE source_file = ?", (sf,)
        ).fetchone()[0]
        conn.execute("UPDATE bruegel_meta SET row_count = ? WHERE source_file = ?", (count, sf))
    conn.commit()

    return inserted


# ---------------------------------------------------------------------------
# Query helpers (for dashboard)
# ---------------------------------------------------------------------------

def get_bruegel_snapshots(conn: sqlite3.Connection) -> list[dict]:
    """Return the most recent value for each Bruegel indicator."""
    cur = conn.execute("""
        SELECT indicator, category, value, date
        FROM bruegel_series
        WHERE id IN (
            SELECT MAX(id) FROM bruegel_series GROUP BY indicator
        )
        ORDER BY category, indicator
    """)
    results = []
    for row in cur.fetchall():
        results.append({
            "indicator": row[0],
            "category": row[1],
            "latest_value": row[2],
            "change": None,
            "unit": "",
            "data_date": row[3],
            "source": "bruegel",
        })
    return results


def get_bruegel_series(conn: sqlite3.Connection, indicator: str, limit: int = 180) -> list[dict]:
    """Return time series for a specific Bruegel indicator."""
    cur = conn.execute(
        "SELECT date, value FROM bruegel_series WHERE indicator = ? ORDER BY date DESC LIMIT ?",
        (indicator, limit),
    )
    return [{"date": row[0], "value": row[1]} for row in cur.fetchall()]


def get_bruegel_indicators(conn: sqlite3.Connection) -> list[dict]:
    """Return distinct indicator/category pairs."""
    cur = conn.execute(
        "SELECT DISTINCT indicator, category FROM bruegel_series ORDER BY category, indicator"
    )
    return [{"indicator": r[0], "category": r[1]} for r in cur.fetchall()]


# ---------------------------------------------------------------------------
# Main fetch orchestrator
# ---------------------------------------------------------------------------

def fetch_all_bruegel(conn: sqlite3.Connection | None = None) -> tuple[list, list, list, int, int]:
    """Download and parse all Bruegel Excel files.
    Returns (all_rows, all_snapshots, provincial_rows, ok_count, fail_count).
    """
    own_conn = conn is None
    if own_conn:
        conn = get_bruegel_db()

    all_rows = []
    all_snapshots = []
    all_provincial = []
    ok = 0
    fail = 0

    for filename in SOURCE_FILES:
        try:
            data = _download_if_modified(filename, conn)
            if data is None:
                # 304 or error — if meta exists it's a skip, otherwise a fail
                row = conn.execute(
                    "SELECT 1 FROM bruegel_meta WHERE source_file = ?", (filename,)
                ).fetchone()
                if row:
                    ok += 1  # already cached
                else:
                    fail += 1
                continue

            if filename in PIVOT_FILES:
                prov_rows = _parse_provincial(data, filename)
                all_provincial.extend(prov_rows)
                log.info(f"  Parsed {filename}: {len(prov_rows)} provincial data points")
            else:
                rows, snaps = _parse_excel(data, filename)
                all_rows.extend(rows)
                all_snapshots.extend(snaps)
                log.info(f"  Parsed {filename}: {len(rows)} data points, {len(snaps)} indicators")
            ok += 1
        except Exception as e:
            log.warning(f"  FAIL {filename}: {type(e).__name__}: {str(e)[:120]}")
            fail += 1

    if own_conn:
        conn.close()

    return all_rows, all_snapshots, all_provincial, ok, fail


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    parser = argparse.ArgumentParser(description="Fetch Bruegel China Economic Database")
    parser.add_argument("--show", action="store_true", help="Show latest stored data")
    parser.add_argument("--json", action="store_true", dest="as_json", help="Output as JSON")
    args = parser.parse_args()

    conn = get_bruegel_db()

    if args.show:
        snapshots = get_bruegel_snapshots(conn)
        if args.as_json:
            print(json.dumps(snapshots, ensure_ascii=False, indent=2))
        else:
            print(f"{'Indicator':<45} {'Value':>12} {'Date':<12} {'Category':<15}")
            print("-" * 90)
            for s in snapshots:
                print(f"{s['indicator']:<45} {s['latest_value']:>12.4f} {s['data_date']:<12} {s['category']:<15}")
            print(f"\nTotal: {len(snapshots)} indicators")
        conn.close()
        return

    log.info("Fetching Bruegel China Economic Database...")
    all_rows, all_snapshots, all_provincial, ok, fail = fetch_all_bruegel(conn)
    inserted = store_bruegel_data(conn, all_rows, all_snapshots)
    prov_inserted = store_provincial_data(conn, all_provincial) if all_provincial else 0
    log.info(f"Done: {ok} OK, {fail} failed, {inserted} data points, {len(all_snapshots)} snapshots, {prov_inserted} provincial")

    total = conn.execute("SELECT COUNT(*) FROM bruegel_series").fetchone()[0]
    prov_total = conn.execute("SELECT COUNT(*) FROM bruegel_provincial").fetchone()[0]
    log.info(f"Total in DB: {total} series + {prov_total} provincial data points")
    conn.close()


if __name__ == "__main__":
    main()
