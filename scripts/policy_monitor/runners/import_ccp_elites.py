"""
Import Sine CPC Elite Leadership Database from Excel into ccp_elites.db.

Usage:
    python -m policy_monitor.runners.import_ccp_elites
    python -m policy_monitor.runners.import_ccp_elites --xlsx data/raw/cpc_elite_leadership.xlsx

Source: https://cpcleadershipdata.pages.dev/CPC_Elite_Leadership_Database.xlsx
"""

import argparse
import logging
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from policy_monitor.ccp_elites import SCHEMA, _DB_DIR, _DB_PATH, get_db

log = logging.getLogger(__name__)

SOURCE_URL = "https://cpcleadershipdata.pages.dev/CPC_Elite_Leadership_Database.xlsx"
DEFAULT_XLSX = Path(__file__).parent.parent.parent.parent / "data" / "raw" / "cpc_elite_leadership.xlsx"


def download_xlsx(dest: Path) -> None:
    log.info(f"Downloading {SOURCE_URL} → {dest}")
    dest.parent.mkdir(parents=True, exist_ok=True)
    urllib.request.urlretrieve(SOURCE_URL, dest)
    log.info(f"Downloaded {dest.stat().st_size / 1024:.0f} KB")


def _str(v) -> str | None:
    return str(v).strip() if v is not None else None


def _int(v) -> int | None:
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def run(xlsx_path: Path = DEFAULT_XLSX, force_download: bool = False) -> None:
    if force_download or not xlsx_path.exists():
        download_xlsx(xlsx_path)

    log.info(f"Loading {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    conn = get_db()

    # --- CC Members ---
    ws_cc = wb["CC Members"]
    cc_rows = []
    for row in ws_cc.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cc_rows.append((
            _str(row[0]),   # congress
            _str(row[1]),   # name
            _str(row[2]),   # name_cn
            _int(row[3]),   # birth_year
            _str(row[4]),   # province
            _str(row[5]),   # is_alternate
            _str(row[6]),   # is_politburo
            _str(row[7]),   # is_psc
            _int(row[8]),   # entry_year
            _int(row[9]),   # exit_year
            _int(row[10]),  # congresses_served
            _str(row[11]),  # expelled
            _str(row[12]),  # expelled_when
            _str(row[13]),  # fate
            _str(row[14]),  # in_previous_cc
        ))

    # --- PB Members ---
    ws_pb = wb["PB Members"]
    pb_rows = []
    for row in ws_pb.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        pb_rows.append((
            _str(row[0]),   # congress
            _str(row[1]),   # name
            _str(row[2]),   # name_cn
            _int(row[3]),   # birth_year
            _str(row[4]),   # is_psc
            _str(row[5]),   # fate
            _str(row[6]),   # in_previous_pb
            _str(row[7]),   # province
            _int(row[8]),   # congresses_served
        ))

    # --- PSC Members ---
    ws_psc = wb["PSC Members"]
    psc_rows = []
    for row in ws_psc.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        psc_rows.append((
            _str(row[0]),   # congress
            _int(row[1]),   # rank
            _str(row[2]),   # name
            _str(row[3]),   # name_cn
            _int(row[4]),   # birth_year
            _str(row[5]),   # province
            _str(row[6]),   # role
            _str(row[7]),   # notes
            _int(row[8]),   # congresses_served
        ))

    # Write to DB (replace all)
    conn.execute("DELETE FROM ccp_cc_members")
    conn.execute("DELETE FROM ccp_pb_members")
    conn.execute("DELETE FROM ccp_psc_members")

    conn.executemany(
        "INSERT INTO ccp_cc_members "
        "(congress,name,name_cn,birth_year,province,is_alternate,is_politburo,is_psc,"
        "entry_year,exit_year,congresses_served,expelled,expelled_when,fate,in_previous_cc) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        cc_rows,
    )
    conn.executemany(
        "INSERT INTO ccp_pb_members "
        "(congress,name,name_cn,birth_year,is_psc,fate,in_previous_pb,province,congresses_served) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        pb_rows,
    )
    conn.executemany(
        "INSERT INTO ccp_psc_members "
        "(congress,rank,name,name_cn,birth_year,province,role,notes,congresses_served) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        psc_rows,
    )

    conn.execute(
        "INSERT INTO ccp_elites_meta (source,source_url,imported_at,cc_rows,pb_rows,psc_rows) "
        "VALUES (?,?,?,?,?,?)",
        (
            "Sine CPC Elite Leadership Database",
            SOURCE_URL,
            datetime.now(timezone.utc).isoformat(),
            len(cc_rows),
            len(pb_rows),
            len(psc_rows),
        ),
    )
    conn.commit()
    conn.close()

    log.info(f"Imported: {len(cc_rows)} CC · {len(pb_rows)} PB · {len(psc_rows)} PSC rows → {_DB_PATH}")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--download", action="store_true", help="Force re-download")
    args = parser.parse_args()
    run(xlsx_path=args.xlsx, force_download=args.download)
